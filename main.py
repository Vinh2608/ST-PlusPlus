from dataset.semi import SemiDataset
from model.semseg.deeplabv2 import DeepLabV2
from model.semseg.deeplabv3plus import DeepLabV3Plus
from model.semseg.pspnet import PSPNet
from utils import count_params, meanIOU, color_map
import random
import argparse
import openpyxl
from copy import deepcopy
import numpy as np
import os
from PIL import Image
import torch
from torch.nn import CrossEntropyLoss, DataParallel
from torch.optim import SGD
from torch.utils.data import DataLoader
from tqdm import tqdm
import matplotlib.pyplot as plt
from loss import consistency_loss
from loss import consistency_loss_
from dataset.transform import crop, hflip, normalize, resize, blur, cutout
from dataset.transform import crop_img, hflip_img, resize_img
from torchvision import transforms
import torch.nn.functional as F
from utils import consistency_weight
import math
import yaml
import logging
import pprint
import segmentation_models_pytorch as smp
from torch.utils.data import ConcatDataset
import albumentations as A
from albumentations.augmentations.domain_adaptation import FDA
import cv2


MODE = None


def parse_args():
    parser = argparse.ArgumentParser(description='ST and ST++ Framework')

    # basic settings
    parser.add_argument('--data-root', type=str, required=True)

    parser.add_argument('--dataset', type=str, choices=[
                        'pascal', 'cityscapes', 'dataset1', 'dataset2', 'lisc', 'raabin'], default='pascal')
    parser.add_argument('--batch-size', type=int, default=16)
    parser.add_argument('--batch-size-consistency', type=int, default=16)
    parser.add_argument('--lr', type=float, default=None)
    parser.add_argument('--epochs', type=int, default=None)
    parser.add_argument('--time', type=int, required=True)
    parser.add_argument('--consistency_training', type=str, required=True)
    parser.add_argument('--warm_up', type=int, default=None)
    parser.add_argument('--crop-size', type=int, default=None)
    parser.add_argument('--backbone', type=str,
                        choices=['resnet50', 'resnet101'], default='resnet50')
    parser.add_argument('--model', type=str, choices=['deeplabv3plus', 'pspnet', 'deeplabv2', 'unet'],
                        default='deeplabv3plus')

    # semi-supervised settings
    parser.add_argument('--labeled-id-path', type=str, required=True)
    parser.add_argument('--addition', type=str, required=True)

    parser.add_argument('--unlabeled-id-path', type=str, required=True)
    parser.add_argument('--pseudo-mask-path', type=str, required=True)
    parser.add_argument('--pseudo-consistency-mask-path',
                        type=str, default=None)
    parser.add_argument('--config', type=str, required=True)

    parser.add_argument('--save-path', type=str, required=True)

    # arguments for ST++
    parser.add_argument('--reliable-id-path', type=str)
    parser.add_argument('--plus', dest='plus', default=False, action='store_true',
                        help='whether to use ST++')

    args = parser.parse_args()
    args.consistency_training = args.consistency_training == 'True'

    return args


def main(args):
    if not os.path.exists(args.save_path):
        os.makedirs(args.save_path)
    if not os.path.exists(args.pseudo_mask_path):
        os.makedirs(args.pseudo_mask_path)
    if args.plus and args.reliable_id_path is None:
        exit('Please specify reliable-id-path in ST++.')
    if not os.path.exists(os.path.join('log', args.dataset, str(args.time), args.model, args.backbone)):
        os.makedirs(os.path.join('log', args.dataset, str(
            args.time), args.model, args.backbone))

    criterion = CrossEntropyLoss(ignore_index=255)
    criterion2 = CrossEntropyLoss(ignore_index=255, reduction='none')

    valset = SemiDataset(args.dataset, args.data_root, 'val', None)
    valloader = DataLoader(valset, batch_size=4 if args.dataset == 'cityscapes' else 1,
                           shuffle=False, pin_memory=True, num_workers=4, drop_last=False)

    cfg = yaml.load(open(args.config, "r"), Loader=yaml.Loader)

    # <====================== Supervised training with labeled images (SupOnly) ======================>
    global MODE
    MODE = 'train'

    print('\n================> Total stage 1/%i: '
          'Supervised training on labeled images (SupOnly)' % (6 if args.plus else 3))

    trainset_labeled = SemiDataset(
        args.dataset, args.data_root, MODE, args.crop_size, args.labeled_id_path)
    if args.consistency_training:
        trainset_unlabeled = SemiDataset(
            args.dataset, args.data_root, 'consistency_training', args.crop_size, None, args.unlabeled_id_path)

        nsample = len(trainset_unlabeled.ids)
        trainset_labeled.ids = trainset_labeled.ids * \
            math.ceil(len(trainset_unlabeled.ids)/len(trainset_labeled.ids))
        trainset_labeled.ids = trainset_labeled.ids[:nsample]
        trainset_unlabeled_loader = DataLoader(
            trainset_unlabeled, batch_size=args.batch_size_consistency, shuffle=False, pin_memory=True, num_workers=4, drop_last=False)
        if args.plus:
            logging.basicConfig(
                filename=f'log/{args.dataset}/{str(args.time)}/{args.model}/{args.backbone}/training_{args.addition}_consistency_st++.log', level=logging.INFO)
        else:
            logging.basicConfig(
                filename=f'log/{args.dataset}/{str(args.time)}/{args.model}/{args.backbone}/training_{args.addition}_consistency_st.log', level=logging.INFO)
    else:
        trainset_labeled.ids = 2 * \
            trainset_labeled.ids if len(
                trainset_labeled.ids) < 200 else trainset_labeled.ids
        if args.plus:
            logging.basicConfig(
                filename=f'log/{args.dataset}/{str(args.time)}/{args.model}/{args.backbone}/training_{args.addition}_non_consistency_st++log', level=logging.INFO)
        else:
            logging.basicConfig(
                filename=f'log/{args.dataset}/{str(args.time)}/{args.model}/{args.backbone}/training_{args.addition}_non_consistency_st.log', level=logging.INFO)

    trainloader_labeled_loader = DataLoader(trainset_labeled, batch_size=args.batch_size, shuffle=True,
                                            pin_memory=True, num_workers=16, drop_last=True)

    all_args = {**cfg, **vars(args), 'ngpus': 1}
    logging.info('{}\n'.format(pprint.pformat(all_args)))

    model, optimizer = init_basic_elems(args)
    print('\nParams: %.1fM' % count_params(model))

    if args.consistency_training:
        best_model, checkpoints = train(model, trainloader_labeled_loader, valloader,
                                        criterion, criterion2, optimizer, args, trainset_unlabeled_loader, cfg)
    else:
        best_model, checkpoints = train(
            model, trainloader_labeled_loader, valloader, criterion, None, optimizer, args)

    """
        ST framework without selective re-training
    """

    if not args.plus:
        # <============================= Pseudo label all unlabeled images =============================>
        print(
            '\n\n\n================> Total stage 2/3: Pseudo labeling all unlabeled images')

        dataset = SemiDataset(args.dataset, args.data_root,
                              'label', None, None, args.unlabeled_id_path)
        dataloader = DataLoader(dataset, batch_size=1, shuffle=False,
                                pin_memory=True, num_workers=4, drop_last=False)

        label(best_model, dataloader, args)

        # <======================== Re-training on labeled and unlabeled images ========================>
        print('\n\n\n================> Total stage 3/3: Re-training on labeled and unlabeled images')

        MODE = 'semi_train'

        trainset = SemiDataset(args.dataset, args.data_root, MODE, args.crop_size,
                               args.labeled_id_path, args.unlabeled_id_path, args.pseudo_mask_path)
        trainloader = DataLoader(trainset, batch_size=args.batch_size, shuffle=True,
                                 pin_memory=True, num_workers=16, drop_last=True)

        model, optimizer = init_basic_elems(args)

        best_performance, _ = train(
            model, trainloader, valloader, criterion, None, optimizer, args, None, None, True)

       # Try to load the existing workbook or create a new one if it doesn't exist
        try:
            workbook = openpyxl.load_workbook(f'{args.dataset}.xlsx')
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        worksheet = workbook.active

        if worksheet.max_row == 1 and worksheet.max_column == 1:
            # Write headers to the worksheet
            headers = ['Experiment Number', 'Model', 'Mode',
                       'Consistency Training', 'Backbone', 'Dataset Name', 'Performance']
            for col_index, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col_index, value=header)

        data = [
            (args.time, args.model, args.plus, args.consistency_training,
             args.backbone, args.dataset, best_performance),
            # Add more rows as needed
        ]

        next_row = worksheet.max_row + 1

        for row_index, (experiment_num, model, plus, consistency_training, backbone, dataset_name, performance) in enumerate(data, start=next_row):
            worksheet.cell(row=row_index, column=1, value=experiment_num)
            worksheet.cell(row=row_index, column=2, value=model)
            worksheet.cell(row=row_index, column=3, value=plus)
            worksheet.cell(row=row_index, column=4, value=consistency_training)
            worksheet.cell(row=row_index, column=5, value=backbone)
            worksheet.cell(row=row_index, column=6, value=dataset_name)
            worksheet.cell(row=row_index, column=7, value=performance)

        workbook.save(f'{args.dataset}.xlsx')
        return

    """
        ST++ framework with selective re-training
    """
    # <===================================== Select Reliable IDs =====================================>
    print('\n\n\n================> Total stage 2/6: Select reliable images for the 1st stage re-training')

    dataset = SemiDataset(args.dataset, args.data_root,
                          'label', None, None, args.unlabeled_id_path)
    dataloader = DataLoader(dataset, batch_size=1, shuffle=False,
                            pin_memory=True, num_workers=4, drop_last=False)

    id_to_reliability = select_reliable(checkpoints, dataloader, args)

    print('\n\n\n================> Total stage 2.5/6: Select reliable images for the 1st stage re-training')

    reliable_image_path = []
    unreliable_image_path = []

    dataset = SemiDataset(args.dataset, args.data_root,
                          'label', None, None, args.unlabeled_id_path)
    dataloader = DataLoader(dataset, batch_size=1, shuffle=False,
                            pin_memory=True, num_workers=4, drop_last=False)
    
    for elem in id_to_reliability[:int(3/4 * len(id_to_reliability))]:
        reliable_image_path.append(elem[0].split()[0])
    for elem in id_to_reliability[int(3/4 * len(id_to_reliability)):]:
        unreliable_image_path.append(elem[0].split()[0])

    fda_transfer(reliable_image_path, unreliable_image_path)

    # <================================ Pseudo label reliable images =================================>
    print('\n\n\n================> Total stage 3/6: Pseudo labeling reliable images')

    cur_unlabeled_id_path = os.path.join(
        args.reliable_id_path, 'reliable_ids.txt')
    dataset = SemiDataset(args.dataset, args.data_root,
                          'label', None, None, cur_unlabeled_id_path)
    dataloader = DataLoader(dataset, batch_size=1, shuffle=False,
                            pin_memory=True, num_workers=4, drop_last=False)

    label(best_model, dataloader, args)

    # <================================== The 1st stage re-training ==================================>
    print('\n\n\n================> Total stage 4/6: The 1st stage re-training on labeled and reliable unlabeled images')

    MODE = 'semi_train'

    trainset = SemiDataset(args.dataset, args.data_root, MODE, args.crop_size,
                           args.labeled_id_path, cur_unlabeled_id_path, args.pseudo_mask_path)
    trainloader = DataLoader(trainset, batch_size=args.batch_size, shuffle=True,
                             pin_memory=True, num_workers=16, drop_last=True)

    model, optimizer = init_basic_elems(args)

    best_model = train(model, trainloader, valloader,
                       criterion, None, optimizer, args)

    # <=============================== Pseudo label unreliable images ================================>
    print('\n\n\n================> Total stage 5/6: Pseudo labeling unreliable images')

    cur_unlabeled_id_path = os.path.join(
        args.reliable_id_path, 'unreliable_ids.txt')
    dataset = SemiDataset(args.dataset, args.data_root,
                          'label', None, None, cur_unlabeled_id_path)
    dataloader = DataLoader(dataset, batch_size=1, shuffle=False,
                            pin_memory=True, num_workers=4, drop_last=False)

    label(best_model, dataloader, args)

    # <================================== The 2nd stage re-training ==================================>
    print('\n\n\n================> Total stage 6/6: The 2nd stage re-training on labeled and all unlabeled images')

    trainset = SemiDataset(args.dataset, args.data_root, MODE, args.crop_size,
                           args.labeled_id_path, args.unlabeled_id_path, args.pseudo_mask_path)
    trainloader = DataLoader(trainset, batch_size=args.batch_size, shuffle=True,
                             pin_memory=True, num_workers=16, drop_last=True)

    model, optimizer = init_basic_elems(args)

    best_performance, _ = train(
        model, trainloader, valloader, criterion, None, optimizer, args, None, None, True)

    try:
        workbook = openpyxl.load_workbook(f'{args.dataset}.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    worksheet = workbook.active

    if worksheet.max_row == 1 and worksheet.max_column == 1:
        # Write headers to the worksheet
        headers = ['Experiment Number', 'Model', 'Mode',
                   'Consistency Training', 'Backbone', 'Dataset Name', 'Performance']
        for col_index, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_index, value=header)

    data = [
        (args.time, args.model, args.plus, args.consistency_training,
         args.backbone, args.dataset, best_performance),
        # Add more rows as needed
    ]

    next_row = worksheet.max_row + 1

    for row_index, (experiment_num, model, plus, consistency_training, backbone, dataset_name, performance) in enumerate(data, start=next_row):
        worksheet.cell(row=row_index, column=1, value=experiment_num)
        worksheet.cell(row=row_index, column=2, value=model)
        worksheet.cell(row=row_index, column=3, value=plus)
        worksheet.cell(row=row_index, column=4, value=consistency_training)
        worksheet.cell(row=row_index, column=5, value=backbone)
        worksheet.cell(row=row_index, column=6, value=dataset_name)
        worksheet.cell(row=row_index, column=7, value=performance)

    workbook.save(f'{args.dataset}.xlsx')


def init_basic_elems(args):
    model_zoo = {'deeplabv3plus': DeepLabV3Plus,
                 'pspnet': PSPNet, 'deeplabv2': DeepLabV2}
    # This is old code
    # model = model_zoo[args.model](args.backbone, 21 if args.dataset == 'pascal' else 19)

    # This is for dataset1 and dataset2
    if args.model != 'unet':
        model = model_zoo[args.model](args.backbone, 3)

        head_lr_multiple = 10.0
        if args.model == 'deeplabv2':
            assert args.backbone == 'resnet101'
            # model.load_state_dict(torch.load('pretrained/deeplabv2_resnet101_coco_pretrained.pth'))
            head_lr_multiple = 1.0

        optimizer = SGD([{'params': model.backbone.parameters(), 'lr': args.lr},
                        {'params': [param for name, param in model.named_parameters()
                                    if 'backbone' not in name],
                        'lr': args.lr * head_lr_multiple}],
                        lr=args.lr, momentum=0.9, weight_decay=1e-4)
    else:
        ENCODER = args.backbone
        ENCODER_WEIGHTS = 'imagenet'
        ACTIVATION = 'softmax2d'

        model = smp.Unet(
            encoder_name=ENCODER,
            encoder_weights=ENCODER_WEIGHTS,
            in_channels=3,
            classes=3,
            activation=ACTIVATION,
        )

        optimizer = torch.optim.Adam([
            dict(params=model.parameters(), lr=0.0001, weight_decay=0.01),
        ])

    model = DataParallel(model).cuda()
    return model, optimizer


def train(model, trainloader, valloader, criterion, criterion2, optimizer, args, consistency_loader=None, cfg=None, final=False):
    global MODE

    iters = 0
    total_iters = len(trainloader) * args.epochs

    iters_per_epoch = len(trainloader)

    previous_best = 0.0

    if MODE == 'train':
        checkpoints = []

    epochs = args.epochs
    if MODE == 'train':
        logging.info(
            f"Training supervised with consistency on unlabeled images")
    else:
        logging.info(f"Semi-training")

    for epoch in range(epochs):
        print("\n==> Epoch %i, learning rate = %.4f\t\t\t\t\t previous best = %.2f" %
              (epoch, optimizer.param_groups[0]["lr"], previous_best))

        model.train()
        total_loss = 0.0
        if consistency_loader != None:
            tbar1 = tqdm(zip(trainloader, consistency_loader))
            for i, ((img, mask), (img_u_w, img_u_s)) in enumerate(tbar1):
                # confidence_threshold=0.95
                img, mask = img.cuda(), mask.cuda()
                img_u_w, img_u_s = img_u_w.cuda(), img_u_s.cuda()

                num_l, num_u = img.shape[0], img_u_w.shape[0]
                pred = model(torch.cat((img, img_u_w)))

                pred_l, pred_w = pred.split([num_l, num_u])

                pred_u_w = pred_w.detach()
                conf_u_w = pred_u_w.softmax(dim=1).max(dim=1)[0]
                mask_u_w = pred_u_w.argmax(dim=1)

                loss_s = criterion(pred_l, mask)

                pred_s = model(img_u_s)

                # loss_unlabeled = softmax_mse_loss(output_strong, output_weak, True,0.9,False)
                # w = consistency_weight(0.002, iters_per_epoch)(epoch, iters)
                # loss_unlabeled = loss_unlabeled * w
                loss_u = criterion2(pred_s, mask_u_w)
                loss_u = loss_u * (conf_u_w >= cfg['conf_thres'])
                loss_u = loss_u.mean()

                # + loss_unlabeled  # Combine the losses if needed
                loss = (loss_s + 0.25 * loss_u)/2
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()

                total_loss += loss.item()

                iters += 1
                logging.info(
                    f"Epoch: {epoch}, Iteration: {iters}, Loss: {loss.item()}, Loss_s: {loss_s.item()}, Loss_u: {loss_u.item()}")

                lr = args.lr * (1 - iters / total_iters) ** 0.9
                if args.model == 'unet':
                    optimizer.param_groups[0]['lr'] = lr
                else:
                    optimizer.param_groups[0]["lr"] = lr
                    optimizer.param_groups[1]["lr"] = lr * \
                        1.0 if args.model == 'deeplabv2' else lr * 10.0

                tbar1.set_description('Loss: %.3f' % (total_loss / (i + 1)))
        else:
            tbar1 = tqdm(trainloader)
            for i, (img, mask) in enumerate(tbar1):
                img, mask = img.cuda(), mask.cuda()

                pred = model(img)

                loss = criterion(pred, mask)

                optimizer.zero_grad()
                loss.backward()
                optimizer.step()

                total_loss += loss.item()

                iters += 1
                logging.info(
                    f"Epoch: {epoch}, Iteration: {iters}, Loss: {loss.item()}")
                lr = args.lr * (1 - iters / total_iters) ** 0.9
                if args.model == 'unet':
                    optimizer.param_groups[0]['lr'] = lr
                else:
                    optimizer.param_groups[0]["lr"] = lr
                    optimizer.param_groups[1]["lr"] = lr * \
                        1.0 if args.model == 'deeplabv2' else lr * 10.0

                tbar1.set_description('Loss: %.3f' % (total_loss / (i + 1)))
        # This is old code
        # metric = meanIOU(num_classes=21 if args.dataset == 'pascal' else 19)
        metric = meanIOU(num_classes=3)
        model.eval()
        tbar = tqdm(valloader)

        with torch.no_grad():
            for img, mask, _ in tbar:
                img = img.cuda()
                pred = model(img)
                pred = torch.argmax(pred, dim=1)

                metric.add_batch(pred.cpu().numpy(), mask.numpy())
                mIOU = metric.evaluate()[-1]

                tbar.set_description('mIOU: %.2f' % (mIOU * 100.0))

        mIOU *= 100.0
        logging.info(f"Validation set mIOU: {mIOU}")
        if mIOU > previous_best:
            if previous_best != 0:
                os.remove(os.path.join(args.save_path, '%s_%s_%.2f.pth' %
                          (args.model, args.backbone, previous_best)))
            previous_best = mIOU
            torch.save(model.module.state_dict(),
                       os.path.join(args.save_path, '%s_%s_%.2f.pth' % (args.model, args.backbone, mIOU)))

            best_model = deepcopy(model)

        if MODE == 'train' and ((epoch + 1) in [epochs // 3, epochs * 2 // 3, epochs]):
            checkpoints.append(deepcopy(model))

    if MODE == 'train':
        return best_model, checkpoints

    if final == False:
        return best_model
    else:
        return previous_best, best_model


def select_reliable(models, dataloader, args):
    if not os.path.exists(args.reliable_id_path):
        os.makedirs(args.reliable_id_path)

    for i in range(len(models)):
        models[i].eval()
    tbar = tqdm(dataloader)

    id_to_reliability = []

    with torch.no_grad():
        for img, mask, id in tbar:
            img = img.cuda()

            preds = []
            for model in models:
                preds.append(torch.argmax(model(img), dim=1).cpu().numpy())

            mIOU = []
            for i in range(len(preds) - 1):
                # metric = meanIOU(num_classes=21 if args.dataset == 'pascal' else 19)
                metric = meanIOU(num_classes=3)
                metric.add_batch(preds[i], preds[-1])
                mIOU.append(metric.evaluate()[-1])

            reliability = sum(mIOU) / len(mIOU)
            id_to_reliability.append((id[0], reliability))

    id_to_reliability.sort(key=lambda elem: elem[1], reverse=True)
    with open(os.path.join(args.reliable_id_path, 'reliable_ids.txt'), 'w') as f:
        for elem in id_to_reliability[:int(3/4 * len(id_to_reliability))]:
            f.write(elem[0] + '\n')
    with open(os.path.join(args.reliable_id_path, 'unreliable_ids.txt'), 'w') as f:
        for elem in id_to_reliability[int(3/4 * len(id_to_reliability)):]:
            f.write(elem[0] + '\n')
    return id_to_reliability


def fda_transfer(reliable_image_paths, unreliable_image_paths, beta_limit=0.1):
    def read_img_pil(file_path):
        with Image.open(file_path) as img:
            return np.array(img.convert('RGB'))

    # Initialize the FDA transformation with the Pillow-based read function
    fda_transform = FDA(
        reference_images=unreliable_image_paths,  # Paths to unreliable images
        beta_limit=beta_limit,
        read_fn=read_img_pil,  # Use the Pillow-based function for reading images
        p=1.0  # Apply the transformation to all images
    )
    
    # Loop over the reliable images and apply the FDA transformation
    for image_path in reliable_image_paths:
        image = read_img_pil(image_path)
        transformed = fda_transform(image=image)['image']
        
        # Save or display the transformed image using PIL
        transformed_path = image_path
        transformed_img = Image.fromarray(transformed)
        transformed_img.save(transformed_path)

def label(model, dataloader, args):
    model.eval()
    tbar = tqdm(dataloader)

    global MODE
    # This is old code
    # metric = meanIOU(num_classes=21 if args.dataset == 'pascal' else 19)
    metric = meanIOU(num_classes=3)

    cmap = color_map(args.dataset)

    with torch.no_grad():
        for img, mask, id in tbar:
            img = img.cuda()
            if args.model != 'unet':
                pred = model(img, True)
            else:
                pred = model(img)
            pred = torch.argmax(pred, dim=1).cpu()
            # # Create a new figure
            # plt.figure(figsize=(10,10))

            # plt.subplot(1, 3, 1)
            # plt.imshow(img.cpu().detach().numpy().squeeze(0).transpose(1,2,0))
            # plt.title('Image')

            # # Plot `mask` on the left
            # plt.subplot(1, 3, 2)
            # plt.imshow(mask.numpy().squeeze(0), cmap='gray')
            # plt.title('Mask')

            # # Plot `pred` on the right
            # plt.subplot(1, 3, 3)
            # plt.imshow(pred.numpy().squeeze(0), cmap='gray')
            # plt.title('Prediction')

            # # Display the figure
            # plt.show()

            metric.add_batch(pred.numpy(), mask.numpy())
            mIOU = metric.evaluate()[-1]

            pred = Image.fromarray(pred.squeeze(
                0).numpy().astype(np.uint8), mode='P')
            pred.putpalette(cmap)
            if args.pseudo_consistency_mask_path is None:
                pred.save('%s/%s' % (args.pseudo_mask_path,
                          os.path.basename(id[0].split(' ')[1])))
            else:
                pred.save('%s/%s' % (args.pseudo_consistency_mask_path,
                          os.path.basename(id[0].split(' ')[1])))

            tbar.set_description('mIOU: %.2f' % (mIOU * 100.0))


if __name__ == '__main__':
    args = parse_args()

    if args.epochs is None:
        args.epochs = {'pascal': 80, 'cityscapes': 240,
                       'dataset1': 100, 'dataset2': 100, 'lisc': 100}[args.dataset]
    if args.lr is None:
        args.lr = {'pascal': 0.001, 'cityscapes': 0.004, 'dataset1': 0.0009,
                   'dataset2': 0.0009, 'lisc': 0.0009}[args.dataset] / 16 * args.batch_size
    if args.warm_up is None:
        args.warm_up = {'pascal': 20, 'cityscapes': 20,
                        'dataset1': 20, 'dataset2': 30, 'lisc': 30}[args.dataset]

    if args.crop_size is None:
        if args.dataset == 'dataset1':
            args.crop_size = {'pascal': 321,
                              'cityscapes': 721, 'dataset1': 128}[args.dataset]
        elif args.dataset == 'dataset2':
            args.crop_size = {'pascal': 321,
                              'cityscapes': 721, 'dataset2': 320}[args.dataset]
        elif args.dataset == 'lisc':
            args.crop_size = {'pascal': 321, 'cityscapes': 721,
                              'dataset2': 320, 'lisc': 128}[args.dataset]

    print()
    print(args)

    main(args)
